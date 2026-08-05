from __future__ import annotations

import csv
import base64
import hashlib
import io
import json
import secrets
import os
import re
import time
import zipfile
import xml.etree.ElementTree as ET
from decimal import Decimal
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Any
from urllib.parse import quote

from fastapi import FastAPI, File, Header, HTTPException, UploadFile
from fastapi.responses import FileResponse, JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, field_validator

ROOT = Path(__file__).resolve().parent
DB_NAME = os.getenv("SQLSERVER_DATABASE", "Maintenance Contract")
DB_SERVER = os.getenv("SQLSERVER_HOST", "localhost")
SESSION_IDLE_SECONDS = 5 * 60
SESSIONS: dict[str, dict[str, Any]] = {}


class SqlCursor:
    def __init__(self, cursor):
        self.cursor = cursor

    def _row(self, row):
        if row is None:
            return None
        return dict(zip((column[0] for column in self.cursor.description), row))

    def fetchone(self):
        return self._row(self.cursor.fetchone())

    def __iter__(self):
        columns = [column[0] for column in self.cursor.description]
        for row in self.cursor:
            yield dict(zip(columns, row))

    @property
    def rowcount(self):
        return self.cursor.rowcount


class SqlConnection:
    def __init__(self, connection):
        self.connection = connection

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc, traceback):
        self.connection.rollback() if exc_type else self.connection.commit()
        self.connection.close()

    def execute(self, sql, params=()):
        return SqlCursor(self.connection.cursor().execute(sql, params))


def db():
    import pyodbc
    installed = set(pyodbc.drivers())
    driver = next((name for name in ('ODBC Driver 18 for SQL Server', 'ODBC Driver 17 for SQL Server') if name in installed), None)
    if not driver:
        raise RuntimeError('Install Microsoft ODBC Driver 18 or 17 for SQL Server')
    connection_string = (
        f"DRIVER={{{driver}}};"
        f"SERVER={DB_SERVER};DATABASE={DB_NAME};"
        "Trusted_Connection=yes;Encrypt=no;TrustServerCertificate=yes;"
    )
    last_error = None
    for attempt in range(1, 13):
        try:
            return SqlConnection(pyodbc.connect(connection_string, timeout=8))
        except pyodbc.Error as error:
            last_error = error
            if attempt < 12:
                time.sleep(5)
    raise RuntimeError(
        f"Could not connect to SQL Server {DB_SERVER} after 12 attempts: {last_error}"
    ) from last_error


def password_hash(value: str) -> str:
    return hashlib.sha256(value.encode()).hexdigest()


def init_db() -> None:
    with db() as con:
        con.execute("""IF OBJECT_ID('properties','U') IS NULL CREATE TABLE properties(
          id INT IDENTITY(1,1) PRIMARY KEY, name NVARCHAR(200) UNIQUE NOT NULL,
          logo NVARCHAR(1000) NOT NULL DEFAULT '', active BIT NOT NULL DEFAULT 0)""")
        con.execute("""IF OBJECT_ID('groups_tbl','U') IS NULL CREATE TABLE groups_tbl(
          id INT IDENTITY(1,1) PRIMARY KEY, name NVARCHAR(200) UNIQUE NOT NULL,
          description NVARCHAR(1000) NOT NULL DEFAULT '')""")
        con.execute("""IF OBJECT_ID('users','U') IS NULL CREATE TABLE users(
          id INT IDENTITY(1,1) PRIMARY KEY, username NVARCHAR(200) UNIQUE NOT NULL,
          password_hash NVARCHAR(64) NOT NULL, role NVARCHAR(30) NOT NULL DEFAULT 'user',
          group_name NVARCHAR(200) NOT NULL DEFAULT '', property_name NVARCHAR(200) NOT NULL DEFAULT '')""")
        for column, definition in (
            ('all_properties', 'BIT NOT NULL DEFAULT 0'),
            ('can_view', 'BIT NOT NULL DEFAULT 1'),
            ('can_create', 'BIT NOT NULL DEFAULT 0'),
            ('can_edit', 'BIT NOT NULL DEFAULT 0'),
            ('can_archive', 'BIT NOT NULL DEFAULT 0'),
            ('can_restore', 'BIT NOT NULL DEFAULT 0'),
            ('can_delete', 'BIT NOT NULL DEFAULT 0'),
            ('can_manage_users', 'BIT NOT NULL DEFAULT 0'),
        ):
            con.execute(f"IF COL_LENGTH('users','{column}') IS NULL ALTER TABLE users ADD {column} {definition}")
        con.execute("""IF OBJECT_ID('user_properties','U') IS NULL CREATE TABLE user_properties(
          user_id INT NOT NULL, property_name NVARCHAR(200) NOT NULL,
          CONSTRAINT PK_user_properties PRIMARY KEY(user_id,property_name))""")
        con.execute("""IF OBJECT_ID('contracts','U') IS NULL CREATE TABLE contracts(
          id INT IDENTITY(1,1) PRIMARY KEY, name NVARCHAR(500) NOT NULL,
          contractor_name NVARCHAR(500) NOT NULL, contractor_phone NVARCHAR(100) NOT NULL DEFAULT '',
          department NVARCHAR(200) NOT NULL DEFAULT '', value DECIMAL(18,2) NOT NULL DEFAULT 0,
          currency NVARCHAR(20) NOT NULL DEFAULT 'JD', tax_percent DECIMAL(8,2) NOT NULL DEFAULT 0,
          start_date DATE NULL, end_date DATE NULL, property_name NVARCHAR(200) NOT NULL DEFAULT '',
          notes NVARCHAR(MAX) NOT NULL DEFAULT '', archived BIT NOT NULL DEFAULT 0,
          created_at DATETIMEOFFSET NOT NULL, updated_at DATETIMEOFFSET NOT NULL)""")
        con.execute("IF COL_LENGTH('contracts','attachment_name') IS NULL ALTER TABLE contracts ADD attachment_name NVARCHAR(500) NULL")
        con.execute("IF COL_LENGTH('contracts','attachment_content_type') IS NULL ALTER TABLE contracts ADD attachment_content_type NVARCHAR(200) NULL")
        con.execute("IF COL_LENGTH('contracts','attachment_data') IS NULL ALTER TABLE contracts ADD attachment_data VARBINARY(MAX) NULL")
        con.execute("IF COL_LENGTH('contracts','attachment_size') IS NULL ALTER TABLE contracts ADD attachment_size BIGINT NULL")
        con.execute("IF COL_LENGTH('contracts','archived_at') IS NULL ALTER TABLE contracts ADD archived_at DATETIMEOFFSET NULL")
        con.execute("IF COL_LENGTH('contracts','archive_exempt') IS NULL ALTER TABLE contracts ADD archive_exempt BIT NOT NULL DEFAULT 0")
        con.execute("IF COL_LENGTH('contracts','created_by') IS NULL ALTER TABLE contracts ADD created_by NVARCHAR(200) NOT NULL DEFAULT 'system'")
        con.execute("IF COL_LENGTH('contracts','updated_by') IS NULL ALTER TABLE contracts ADD updated_by NVARCHAR(200) NOT NULL DEFAULT 'system'")
        con.execute("IF COL_LENGTH('contracts','archived_by') IS NULL ALTER TABLE contracts ADD archived_by NVARCHAR(200) NULL")
        con.execute("UPDATE contracts SET archived_at=COALESCE(updated_at,created_at) WHERE archived=1 AND archived_at IS NULL")
        con.execute("""IF OBJECT_ID('audit_logs','U') IS NULL CREATE TABLE audit_logs(
          id INT IDENTITY(1,1) PRIMARY KEY, username NVARCHAR(200) NOT NULL,
          action NVARCHAR(100) NOT NULL, entity_type NVARCHAR(100) NOT NULL,
          entity_id INT NULL, details NVARCHAR(1000) NOT NULL DEFAULT '',
          property_name NVARCHAR(200) NOT NULL DEFAULT '', created_at DATETIMEOFFSET NOT NULL)""")
        con.execute("""IF OBJECT_ID('departments','U') IS NULL CREATE TABLE departments(
          id INT IDENTITY(1,1) PRIMARY KEY, name NVARCHAR(200) NOT NULL UNIQUE,
          description NVARCHAR(500) NOT NULL DEFAULT '', active BIT NOT NULL DEFAULT 1,
          created_at DATETIMEOFFSET NOT NULL)""")
        con.execute("""IF OBJECT_ID('currencies','U') IS NULL CREATE TABLE currencies(
          code NVARCHAR(20) PRIMARY KEY, name NVARCHAR(100) NOT NULL,
          symbol NVARCHAR(20) NOT NULL DEFAULT '', active BIT NOT NULL DEFAULT 1,
          created_at DATETIMEOFFSET NOT NULL)""")
        con.execute("""IF OBJECT_ID('system_settings','U') IS NULL CREATE TABLE system_settings(
          setting_key NVARCHAR(100) PRIMARY KEY, setting_value NVARCHAR(MAX) NOT NULL DEFAULT '',
          updated_at DATETIMEOFFSET NOT NULL)""")
        con.execute("""IF OBJECT_ID('property_logos','U') IS NULL CREATE TABLE property_logos(
          id INT IDENTITY(1,1) PRIMARY KEY, property_name NVARCHAR(200) NOT NULL UNIQUE,
          logo_name NVARCHAR(500) NOT NULL, content_type NVARCHAR(100) NOT NULL,
          logo_data VARBINARY(MAX) NOT NULL, updated_at DATETIMEOFFSET NOT NULL,
          updated_by NVARCHAR(200) NOT NULL DEFAULT '')""")
        con.execute("""IF OBJECT_ID('database_backups','U') IS NULL CREATE TABLE database_backups(
          id INT IDENTITY(1,1) PRIMARY KEY, backup_name NVARCHAR(500) NOT NULL,
          backup_data VARBINARY(MAX) NULL, backup_size BIGINT NOT NULL,
          backup_path NVARCHAR(1000) NULL, created_at DATETIMEOFFSET NOT NULL, created_by NVARCHAR(200) NOT NULL)""")
        con.execute("ALTER TABLE database_backups ALTER COLUMN backup_data VARBINARY(MAX) NULL")
        con.execute("IF COL_LENGTH('database_backups','backup_path') IS NULL ALTER TABLE database_backups ADD backup_path NVARCHAR(1000) NULL")
        con.execute("""IF OBJECT_ID('database_backup_chunks','U') IS NULL CREATE TABLE database_backup_chunks(
          id INT IDENTITY(1,1) PRIMARY KEY, backup_id INT NOT NULL,
          chunk_index INT NOT NULL, chunk_data VARBINARY(MAX) NOT NULL)""")
        con.execute("""IF OBJECT_ID('schema_migrations','U') IS NULL CREATE TABLE schema_migrations(
          version INT PRIMARY KEY, description NVARCHAR(500) NOT NULL,
          applied_at DATETIMEOFFSET NOT NULL)""")
        con.execute("DELETE c FROM database_backup_chunks c LEFT JOIN database_backups b ON b.id=c.backup_id WHERE b.id IS NULL")
        con.execute("WITH duplicate_chunks AS (SELECT id,ROW_NUMBER() OVER(PARTITION BY backup_id,chunk_index ORDER BY id) AS row_number FROM database_backup_chunks) DELETE FROM duplicate_chunks WHERE row_number>1")
        con.execute("IF NOT EXISTS(SELECT 1 FROM sys.foreign_keys WHERE name='FK_database_backup_chunks_backup') ALTER TABLE database_backup_chunks ADD CONSTRAINT FK_database_backup_chunks_backup FOREIGN KEY(backup_id) REFERENCES database_backups(id) ON DELETE CASCADE")
        con.execute("IF NOT EXISTS(SELECT 1 FROM sys.indexes WHERE name='UX_database_backup_chunks_order' AND object_id=OBJECT_ID('database_backup_chunks')) CREATE UNIQUE INDEX UX_database_backup_chunks_order ON database_backup_chunks(backup_id,chunk_index)")
        con.execute("IF NOT EXISTS(SELECT 1 FROM sys.indexes WHERE name='IX_contracts_scope_status' AND object_id=OBJECT_ID('contracts')) CREATE INDEX IX_contracts_scope_status ON contracts(property_name,archived,end_date) INCLUDE(name,department,value,currency)")
        con.execute("IF NOT EXISTS(SELECT 1 FROM sys.indexes WHERE name='IX_audit_logs_created' AND object_id=OBJECT_ID('audit_logs')) CREATE INDEX IX_audit_logs_created ON audit_logs(created_at DESC,property_name)")
        con.execute("IF NOT EXISTS(SELECT 1 FROM sys.foreign_keys WHERE name='FK_user_properties_user') ALTER TABLE user_properties ADD CONSTRAINT FK_user_properties_user FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE")
        con.execute("IF NOT EXISTS(SELECT 1 FROM sys.indexes WHERE name='IX_user_properties_property' AND object_id=OBJECT_ID('user_properties')) CREATE INDEX IX_user_properties_property ON user_properties(property_name,user_id)")
        con.execute("IF NOT EXISTS(SELECT 1 FROM sys.indexes WHERE name='IX_database_backups_created' AND object_id=OBJECT_ID('database_backups')) CREATE INDEX IX_database_backups_created ON database_backups(created_at DESC)")
        default_backup_directory=str((ROOT/'backups').resolve()); now_setting=datetime.now(timezone.utc).isoformat()
        con.execute("IF NOT EXISTS(SELECT 1 FROM system_settings WHERE setting_key='backup_directory') INSERT INTO system_settings(setting_key,setting_value,updated_at) VALUES('backup_directory',?,?)",(default_backup_directory,now_setting))
        now = datetime.now(timezone.utc).isoformat()
        con.execute("INSERT INTO departments(name,created_at) SELECT DISTINCT department,? FROM contracts WHERE department<>'' AND NOT EXISTS(SELECT 1 FROM departments d WHERE d.name=contracts.department)",(now,))
        for code,name,symbol in (('JD','Jordanian Dinar','JD'),('USD','US Dollar','$'),('SAR','Saudi Riyal','SAR'),('EUR','Euro','€')):
            con.execute('IF NOT EXISTS(SELECT 1 FROM currencies WHERE code=?) INSERT INTO currencies(code,name,symbol,created_at) VALUES(?,?,?,?)',(code,code,name,symbol,now))
        con.execute("INSERT INTO currencies(code,name,symbol,created_at) SELECT DISTINCT currency,currency,currency,? FROM contracts WHERE currency<>'' AND NOT EXISTS(SELECT 1 FROM currencies c WHERE c.code=contracts.currency)",(now,))
        con.execute("IF COL_LENGTH('contracts','department_id') IS NULL ALTER TABLE contracts ADD department_id INT NULL")
        con.execute("IF COL_LENGTH('contracts','currency_code') IS NULL ALTER TABLE contracts ADD currency_code NVARCHAR(20) NULL")
        con.execute("UPDATE c SET department_id=d.id FROM contracts c JOIN departments d ON d.name=c.department WHERE c.department_id IS NULL")
        con.execute("UPDATE contracts SET currency_code=currency WHERE currency_code IS NULL AND currency<>''")
        con.execute("IF NOT EXISTS(SELECT 1 FROM sys.indexes WHERE name='IX_contracts_department' AND object_id=OBJECT_ID('contracts')) CREATE INDEX IX_contracts_department ON contracts(department_id,currency_code)")
        con.execute("""EXEC(N'CREATE OR ALTER VIEW dbo.vw_contract_management AS
          SELECT c.id,c.name,c.contractor_name,c.contractor_phone,c.department,c.value,c.currency,
          c.tax_percent,c.start_date,c.end_date,c.property_name,c.archived,c.archived_at,
          CASE WHEN c.attachment_data IS NULL THEN CAST(0 AS BIT) ELSE CAST(1 AS BIT) END AS has_attachment,
          CASE WHEN c.end_date IS NULL THEN N''active'' WHEN c.end_date<CAST(GETDATE() AS DATE) THEN N''expired''
               WHEN c.end_date<=DATEADD(DAY,30,CAST(GETDATE() AS DATE)) THEN N''expiring_soon'' ELSE N''active'' END AS contract_status,
          c.created_at,c.updated_at FROM dbo.contracts c')""")
        con.execute("IF NOT EXISTS(SELECT 1 FROM schema_migrations WHERE version=1) INSERT INTO schema_migrations(version,description,applied_at) VALUES(1,'Complete SQL persistence schema',?)",(now,))
        con.execute("IF NOT EXISTS(SELECT 1 FROM properties WHERE name=?) INSERT INTO properties(name,active) VALUES(?,1)", ('Hilton Amman','Hilton Amman'))
        con.execute("IF NOT EXISTS(SELECT 1 FROM groups_tbl WHERE name=?) INSERT INTO groups_tbl(name,description) VALUES(?,?)", ('Administrators','Administrators','System administrators'))
        con.execute("IF NOT EXISTS(SELECT 1 FROM users WHERE username=?) INSERT INTO users(username,password_hash,role,group_name,property_name) VALUES(?,?,?,?,?)",
                    ('admin','admin', password_hash('Admin@123'), 'admin', 'Administrators', 'Hilton Amman'))
        if not con.execute("SELECT version FROM schema_migrations WHERE version=2").fetchone():
            con.execute("UPDATE users SET all_properties=1,can_view=1,can_create=1,can_edit=1,can_archive=1,can_restore=1,can_delete=1,can_manage_users=1 WHERE role='admin'")
            con.execute("INSERT INTO user_properties(user_id,property_name) SELECT id,property_name FROM users u WHERE property_name<>'' AND NOT EXISTS(SELECT 1 FROM user_properties up WHERE up.user_id=u.id AND up.property_name=u.property_name)")
            con.execute("INSERT INTO schema_migrations(version,description,applied_at) VALUES(2,'Role permissions, multiple properties, and contract ownership audit',?)",(now,))


init_db()
app = FastAPI(title="Maintenance Contracts System")


def auth(authorization: str | None) -> dict[str, Any]:
    token = (authorization or '').removeprefix('Bearer ').strip()
    session = SESSIONS.get(token)
    now = time.monotonic()
    if session and now - float(session.get('last_activity', now)) >= SESSION_IDLE_SECONDS:
        SESSIONS.pop(token, None)
        session = None
    if not session:
        raise HTTPException(401, 'Unauthorized')
    session['last_activity'] = now
    with db() as con:
        user = con.execute('''SELECT id,username,role,group_name,property_name,all_properties,
            can_view,can_create,can_edit,can_archive,can_restore,can_delete,can_manage_users
            FROM users WHERE username=?''',(session['username'],)).fetchone()
    if not user:
        raise HTTPException(401, 'Unauthorized')
    result = dict(user)
    result['property_name'] = session['property_name']
    with db() as con:
        result['property_names'] = [x['property_name'] for x in con.execute('SELECT property_name FROM user_properties WHERE user_id=? ORDER BY property_name',(result['id'],))]
    return result


def require_permission(user: dict[str, Any], permission: str) -> None:
    if not bool(user.get(permission)):
        raise HTTPException(403, 'You do not have permission to perform this action')


def write_audit(user: dict[str, Any], action: str, entity_type: str, entity_id: int | None = None, details: str = '', property_name: str = '') -> None:
    with db() as con:
        con.execute('INSERT INTO audit_logs(username,action,entity_type,entity_id,details,property_name,created_at) VALUES(?,?,?,?,?,?,?)',
                    (user.get('username','system'),action,entity_type,entity_id,details[:1000],property_name or user.get('property_name',''),datetime.now(timezone.utc).isoformat()))


class Login(BaseModel):
    username: str
    password: str
    property_name: str = ''


class Contract(BaseModel):
    name: str
    contractor_name: str
    contractor_phone: str = ''
    department: str = ''
    value: float = 0
    currency: str = 'JD'
    tax_percent: float = 0
    start_date: str = ''
    end_date: str = ''
    property_name: str = ''
    notes: str = ''

    @field_validator('name', 'contractor_name')
    @classmethod
    def required_text(cls, value: str) -> str:
        value = value.strip()
        if not value:
            raise ValueError('This field is required')
        return value

    @field_validator('value', 'tax_percent')
    @classmethod
    def non_negative_number(cls, value: float) -> float:
        if value < 0:
            raise ValueError('Value cannot be negative')
        return value

    @field_validator('start_date', 'end_date')
    @classmethod
    def valid_date(cls, value: str) -> str:
        if value:
            date.fromisoformat(value)
        return value


class UserPayload(BaseModel):
    username: str
    password: str = ''
    role: str = 'user'
    group_name: str = ''
    property_name: str = ''
    property_names: list[str] = []
    all_properties: bool = False
    can_view: bool = True
    can_create: bool = False
    can_edit: bool = False
    can_archive: bool = False
    can_restore: bool = False
    can_delete: bool = False
    can_manage_users: bool = False


class NamedPayload(BaseModel):
    name: str
    description: str = ''
    logo: str = ''


class CurrencyPayload(BaseModel):
    code: str
    name: str
    symbol: str = ''


class ReportExportPayload(BaseModel):
    title: str = 'Management Report'
    filters: str = ''
    property_name: str = ''
    rows: list[dict[str, Any]] = []


class BackupSettingsPayload(BaseModel):
    backup_directory: str


@app.post('/api/auth/login')
def login(payload: Login):
    with db() as con:
        row = con.execute('SELECT * FROM users WHERE username=?',(payload.username,)).fetchone()
        selected_property = con.execute('SELECT name FROM properties WHERE name=?',(payload.property_name,)).fetchone() if payload.property_name else None
        assigned = con.execute('SELECT 1 AS ok FROM user_properties WHERE user_id=? AND property_name=?',(row['id'],payload.property_name)).fetchone() if row and payload.property_name else None
    if not row or row['password_hash'] != password_hash(payload.password):
        raise HTTPException(401, 'Invalid username or password')
    if not row['all_properties'] and not selected_property:
        raise HTTPException(400, 'Please select a valid property')
    if not row['all_properties'] and not assigned:
        raise HTTPException(403, 'This user is not assigned to the selected property')
    token = secrets.token_urlsafe(32)
    selected_name = payload.property_name if selected_property else ''
    SESSIONS[token] = {'username': row['username'], 'property_name': selected_name, 'last_activity': time.monotonic()}
    write_audit(dict(row),'login','session',None,'Successful login',selected_name)
    return {'token': token, 'user': {'username': row['username'], 'role': row['role'], 'property_name': selected_name}}


@app.post('/api/auth/logout')
def logout(authorization: str | None = Header(None)):
    token = (authorization or '').removeprefix('Bearer ').strip()
    SESSIONS.pop(token, None)
    return {'success': True}


@app.get('/api/auth/properties')
def login_properties():
    with db() as con:
        return [dict(x) for x in con.execute('SELECT name,logo FROM properties ORDER BY name')]


@app.get('/api/me')
def me(authorization: str | None = Header(None)):
    return dict(auth(authorization))


def status_for(end_date: str) -> str:
    try:
        parsed = end_date if isinstance(end_date, date) else date.fromisoformat(str(end_date))
        days = (parsed - date.today()).days
    except Exception: return 'active'
    return 'expired' if days < 0 else ('expiring_soon' if days <= 30 else 'active')


def archive_expired_contracts() -> None:
    """Archive contracts after their final valid day has passed."""
    with db() as con:
        con.execute('''UPDATE contracts
            SET archived=1, archived_at=?, archived_by='system', updated_at=?, updated_by='system'
            WHERE archived=0 AND archive_exempt=0 AND end_date IS NOT NULL AND end_date < ?''',
            (datetime.now(timezone.utc).isoformat(), datetime.now(timezone.utc).isoformat(), date.today().isoformat()))


@app.get('/api/contracts')
def list_contracts(include_archived: bool = False, authorization: str | None = Header(None)):
    user = auth(authorization)
    require_permission(user, 'can_view')
    archive_expired_contracts()
    sql, args = '''SELECT id,name,contractor_name,contractor_phone,department,value,currency,
        tax_percent,start_date,end_date,property_name,notes,archived,
        CONVERT(NVARCHAR(40),archived_at,127) AS archived_at,
        attachment_name,attachment_content_type,attachment_size,
        CASE WHEN attachment_data IS NULL THEN CAST(0 AS BIT) ELSE CAST(1 AS BIT) END AS has_attachment,
        CONVERT(NVARCHAR(40),created_at,127) AS created_at,
        CONVERT(NVARCHAR(40),updated_at,127) AS updated_at,
        created_by,updated_by,archived_by
        FROM contracts WHERE archived=?''', [1 if include_archived else 0]
    if user['property_name']:
        sql += ' AND property_name=?'; args.append(user['property_name'])
    sql += ' ORDER BY end_date'
    with db() as con: rows = [dict(x) for x in con.execute(sql,args)]
    for row in rows: row['status'] = status_for(row['end_date'])
    return rows


@app.post('/api/contracts')
def add_contract(payload: Contract, authorization: str | None = Header(None)):
    user = auth(authorization)
    require_permission(user, 'can_create')
    data = payload.model_dump()
    if user['property_name']:
        data['property_name'] = user['property_name']
    if not data['property_name']:
        raise HTTPException(400, 'Please select a property for this contract')
    with db() as con:
        property_exists = con.execute('SELECT id FROM properties WHERE name=?',(data['property_name'],)).fetchone()
    if not property_exists:
        raise HTTPException(400, 'The selected property does not exist')
    with db() as con:
        if not con.execute('SELECT id FROM departments WHERE name=? AND active=1',(data['department'],)).fetchone(): raise HTTPException(400,'Please select a valid department')
        if not con.execute('SELECT code FROM currencies WHERE code=? AND active=1',(data['currency'],)).fetchone(): raise HTTPException(400,'Please select a valid currency')
    data['start_date'] = data['start_date'] or None
    data['end_date'] = data['end_date'] or None
    now = datetime.now(timezone.utc).isoformat()
    with db() as con:
        cur=con.execute('''INSERT INTO contracts(name,contractor_name,contractor_phone,department,value,currency,tax_percent,start_date,end_date,property_name,notes,created_at,updated_at,created_by,updated_by) OUTPUT INSERTED.id VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',(
            data['name'],data['contractor_name'],data['contractor_phone'],data['department'],data['value'],data['currency'],data['tax_percent'],data['start_date'],data['end_date'],data['property_name'],data['notes'],now,now,user['username'],user['username']))
        contract_id=cur.fetchone()['id']
        con.execute('UPDATE contracts SET department_id=(SELECT id FROM departments WHERE name=?),currency_code=? WHERE id=?',(data['department'],data['currency'],contract_id))
    write_audit(user,'created','contract',contract_id,data['name'],data['property_name'])
    return {'success':True,'id':contract_id,'property_name':data['property_name']}


def contract_access(user: dict[str, Any], contract_id: int) -> dict[str, Any]:
    with db() as con:
        contract = con.execute('SELECT id,property_name FROM contracts WHERE id=?',(contract_id,)).fetchone()
    if not contract:
        raise HTTPException(404, 'Contract not found')
    if user['property_name'] and contract['property_name'] != user['property_name']:
        raise HTTPException(403, 'You cannot modify a contract from another property')
    if not user['property_name'] and not user['all_properties'] and contract['property_name'] not in user.get('property_names',[]):
        raise HTTPException(403, 'You cannot access a contract from another property')
    return contract


@app.put('/api/contracts/{contract_id}')
def edit_contract(contract_id: int, payload: Contract, authorization: str | None = Header(None)):
    user=auth(authorization); require_permission(user,'can_edit'); contract_access(user,contract_id); data=payload.model_dump(); now=datetime.now(timezone.utc).isoformat()
    if user['property_name']:
        data['property_name']=user['property_name']
    if not data['property_name']:
        raise HTTPException(400, 'Please select a property for this contract')
    with db() as con:
        if not con.execute('SELECT id FROM properties WHERE name=?',(data['property_name'],)).fetchone():
            raise HTTPException(400, 'The selected property does not exist')
        if not con.execute('SELECT id FROM departments WHERE name=? AND active=1',(data['department'],)).fetchone(): raise HTTPException(400,'Please select a valid department')
        if not con.execute('SELECT code FROM currencies WHERE code=? AND active=1',(data['currency'],)).fetchone(): raise HTTPException(400,'Please select a valid currency')
    data['start_date']=data['start_date'] or None; data['end_date']=data['end_date'] or None
    with db() as con:
        old=con.execute('SELECT name,contractor_name,value,start_date,end_date,property_name FROM contracts WHERE id=?',(contract_id,)).fetchone()
        con.execute('''UPDATE contracts SET name=?,contractor_name=?,contractor_phone=?,department=?,value=?,currency=?,tax_percent=?,start_date=?,end_date=?,property_name=?,notes=?,archive_exempt=0,updated_at=?,updated_by=? WHERE id=?''',(
            data['name'],data['contractor_name'],data['contractor_phone'],data['department'],data['value'],data['currency'],data['tax_percent'],data['start_date'],data['end_date'],data['property_name'],data['notes'],now,user['username'],contract_id))
        con.execute('UPDATE contracts SET department_id=(SELECT id FROM departments WHERE name=?),currency_code=? WHERE id=?',(data['department'],data['currency'],contract_id))
    write_audit(user,'updated','contract',contract_id,json.dumps({'before':old,'after':{k:data[k] for k in ('name','contractor_name','value','start_date','end_date','property_name')}},default=str),data['property_name'])
    return {'success':True}


MAX_ATTACHMENT_SIZE = 25 * 1024 * 1024


def contract_document_text(filename: str, content_type: str, data: bytes) -> str:
    suffix=Path(filename).suffix.lower()
    try:
        if suffix=='.pdf' or content_type=='application/pdf':
            from pypdf import PdfReader
            return '\n'.join((page.extract_text() or '') for page in PdfReader(io.BytesIO(data)).pages)
        if suffix=='.docx':
            with zipfile.ZipFile(io.BytesIO(data)) as archive:
                root=ET.fromstring(archive.read('word/document.xml'))
            return '\n'.join(''.join(node.itertext()) for node in root.iter() if node.tag.endswith('}p'))
        if suffix in {'.txt','.csv'} or content_type.startswith('text/'):
            return data.decode('utf-8-sig',errors='replace')
    except Exception as exc:
        raise HTTPException(400,f'Could not read the selected contract document: {exc}')
    raise HTTPException(400,'Automatic reading supports searchable PDF, DOCX, and text files. Scanned images require OCR before upload.')


def first_document_value(text: str, labels: list[str]) -> str:
    for label in labels:
        match=re.search(rf'(?:{label})\s*[:\-–]?\s*([^\n\r]{{2,180}})',text,re.IGNORECASE)
        if match:
            return match.group(1).strip(' .:-–')
    return ''


def normalize_document_date(value: str) -> str:
    value=(value or '').translate(str.maketrans('٠١٢٣٤٥٦٧٨٩','0123456789'))
    match=re.search(r'(\d{1,4})[./\-](\d{1,2})[./\-](\d{1,4})',value)
    if match:
        a,b,c=map(int,match.groups()); year,month,day=((a,b,c) if a>1900 else (c,b,a))
        try:
            return date(year,month,day).isoformat()
        except ValueError:
            pass
    month_date=re.search(r'([A-Za-z]+)\s+(\d{1,2})(?:st|nd|rd|th)?[,]?\s+(\d{4})|(\d{1,2})(?:st|nd|rd|th)?\s+([A-Za-z]+)[,]?\s+(\d{4})',value,re.IGNORECASE)
    if month_date:
        candidate=' '.join(x for x in month_date.groups() if x)
        for pattern in ('%B %d %Y','%b %d %Y','%d %B %Y','%d %b %Y'):
            try:
                return datetime.strptime(candidate,pattern).date().isoformat()
            except ValueError:
                continue
    return ''


def labelled_document_date(text: str, labels: list[str]) -> str:
    for label in labels:
        match=re.search(label,text,re.IGNORECASE)
        if match:
            parsed=normalize_document_date(text[match.end():match.end()+160])
            if parsed:
                return parsed
    return ''


def document_dates(text: str) -> list[str]:
    candidates=re.findall(r'\d{1,4}[./\-]\d{1,2}[./\-]\d{1,4}|(?:[A-Za-z]+\s+\d{1,2}(?:st|nd|rd|th)?[,]?\s+\d{4})|(?:\d{1,2}(?:st|nd|rd|th)?\s+[A-Za-z]+[,]?\s+\d{4})',text,re.IGNORECASE)
    result=[]
    for candidate in candidates:
        parsed=normalize_document_date(candidate)
        if parsed and parsed not in result:
            result.append(parsed)
    return result


def document_company_phone(text: str) -> str:
    labels=[r'(?:company|contractor|supplier|vendor)?\s*(?:phone|telephone|mobile)(?:\s*(?:number|no\.))?', '\u0631\u0642\u0645[ \t\r\n]*\u0647\u0627\u062a\u0641', '\u0627\u0644\u0647\u0627\u062a\u0641', '\u062c\u0648\u0627\u0644']
    areas=[]
    for label in labels:
        match=re.search(label,text,re.IGNORECASE)
        if match:
            areas.append(text[match.end():match.end()+90])
    areas.append(text)
    for area in areas:
        for match in re.finditer(r'(?<!\d)(?:\+?\d[\d\s().\-]{5,}\d)(?!\d)',area):
            phone=re.sub(r'\s+',' ',match.group(0)).strip(' .-')
            digits=re.sub(r'\D','',phone)
            if 7<=len(digits)<=15 and not re.fullmatch(r'\d{1,4}[./\-]\d{1,2}[./\-]\d{1,4}',phone):
                return phone
    return ''


@app.post('/api/contracts/extract-document')
async def extract_contract_document_enhanced(file: UploadFile = File(...), authorization: str | None = Header(None)):
    auth(authorization)
    data=await file.read(MAX_ATTACHMENT_SIZE+1)
    if not data or len(data)>MAX_ATTACHMENT_SIZE:
        raise HTTPException(400,'Select a contract document between 1 byte and 25 MB')
    text=contract_document_text(Path(file.filename or 'contract').name,file.content_type or '',data)
    compact=re.sub(r'[ \t]+',' ',text)
    start_date=labelled_document_date(compact,[r'contract\s+start\s+date',r'start\s+date',r'effective\s+date','\u062a\u0627\u0631\u064a\u062e[ \t\r\n]*(?:\u0628\u062f\u0621|\u0627\u0644\u0628\u062f\u0627\u064a\u0629)'])
    end_date=labelled_document_date(compact,[r'contract\s+end\s+date',r'end\s+date',r'expir(?:y|ation)\s+date','\u062a\u0627\u0631\u064a\u062e[ \t\r\n]*(?:\u0627\u0646\u062a\u0647\u0627\u0621|\u0627\u0644\u0646\u0647\u0627\u064a\u0629)'])
    dates=document_dates(compact)
    if not start_date and dates: start_date=dates[0]
    if not end_date and len(dates)>1: end_date=dates[1]
    value_raw=first_document_value(compact,[r'contract value',r'total amount','\u0642\u064a\u0645\u0629[ \t\r\n]*\u0627\u0644\u0639\u0642\u062f'])
    tax_raw=first_document_value(compact,[r'tax(?: percentage)?',r'vat','\u0627\u0644\u0636\u0631\u064a\u0628\u0629'])
    currency_match=re.search(r'\b(JOD|JD|USD|EUR|SAR|AED|QAR|BHD|KWD|OMR|GBP)\b',compact,re.IGNORECASE)
    number_match=re.search(r'[-+]?\d[\d,]*(?:\.\d+)?',value_raw)
    tax_match=re.search(r'\d+(?:\.\d+)?',tax_raw)
    result={'name':first_document_value(compact,[r'contract (?:name|title|subject)',r'agreement (?:name|title)','\u0627\u0633\u0645[ \t\r\n]*\u0627\u0644\u0639\u0642\u062f','\u0645\u0648\u0636\u0648\u0639[ \t\r\n]*\u0627\u0644\u0639\u0642\u062f']),'contractor_name':first_document_value(compact,[r'contractor(?: name)?',r'supplier(?: name)?',r'vendor(?: name)?','\u0627\u0633\u0645[ \t\r\n]*\u0627\u0644\u0645\u0642\u0627\u0648\u0644','\u0627\u0633\u0645[ \t\r\n]*\u0627\u0644\u0645\u0648\u0631\u062f']),'contractor_phone':document_company_phone(compact),'value':float(number_match.group(0).replace(',','')) if number_match else None,'currency':currency_match.group(1).upper() if currency_match else '','tax_percent':float(tax_match.group(0)) if tax_match else None,'start_date':start_date,'end_date':end_date,'notes':first_document_value(compact,[r'(?:scope of work|description|notes)','\u0646\u0637\u0627\u0642[ \t\r\n]*\u0627\u0644\u0639\u0645\u0644','\u0645\u0644\u0627\u062d\u0638\u0627\u062a'])}
    result['extracted_fields']=[key for key,value in result.items() if value not in ('',None)]
    return result


@app.post('/api/contracts/extract-document-legacy',include_in_schema=False)
async def extract_contract_document(file: UploadFile = File(...), authorization: str | None = Header(None)):
    auth(authorization)
    data=await file.read(MAX_ATTACHMENT_SIZE+1)
    if not data or len(data)>MAX_ATTACHMENT_SIZE:
        raise HTTPException(400,'Select a contract document between 1 byte and 25 MB')
    text=contract_document_text(Path(file.filename or 'contract').name,file.content_type or '',data)
    compact=re.sub(r'[ \t]+',' ',text)
    start_raw=first_document_value(compact,[r'contract start date',r'start date',r'effective date',r'تاريخ بدء العقد',r'تاريخ البداية'])
    end_raw=first_document_value(compact,[r'contract end date',r'end date',r'expiry date',r'expiration date',r'تاريخ انتهاء العقد',r'تاريخ النهاية'])
    value_raw=first_document_value(compact,[r'contract value',r'total amount',r'قيمة العقد',r'القيمة الإجمالية'])
    tax_raw=first_document_value(compact,[r'tax(?: percentage)?',r'vat',r'الضريبة'])
    currency_match=re.search(r'\b(JOD|JD|USD|EUR|SAR|AED|QAR|BHD|KWD|OMR|GBP)\b',compact,re.IGNORECASE)
    number_match=re.search(r'[-+]?\d[\d,]*(?:\.\d+)?',value_raw)
    tax_match=re.search(r'\d+(?:\.\d+)?',tax_raw)
    result={'name':first_document_value(compact,[r'contract (?:name|title|subject)',r'agreement (?:name|title)',r'اسم العقد',r'موضوع العقد']),'contractor_name':first_document_value(compact,[r'contractor(?: name)?',r'supplier(?: name)?',r'vendor(?: name)?',r'اسم المقاول',r'اسم المورد']),'contractor_phone':first_document_value(compact,[r'(?:phone|telephone|mobile)(?: number)?',r'رقم الهاتف',r'هاتف']),'value':float(number_match.group(0).replace(',','')) if number_match else None,'currency':currency_match.group(1).upper() if currency_match else '','tax_percent':float(tax_match.group(0)) if tax_match else None,'start_date':normalize_document_date(start_raw),'end_date':normalize_document_date(end_raw),'notes':first_document_value(compact,[r'(?:scope of work|description|notes)',r'نطاق العمل',r'ملاحظات'])}
    result['extracted_fields']=[key for key,value in result.items() if value not in ('',None)]
    return result


@app.post('/api/contracts/{contract_id}/attachment')
async def upload_contract_attachment(contract_id: int, file: UploadFile = File(...), authorization: str | None = Header(None)):
    user = auth(authorization)
    require_permission(user,'can_edit')
    contract_access(user, contract_id)
    data = await file.read(MAX_ATTACHMENT_SIZE + 1)
    if not data:
        raise HTTPException(400, 'Please select a non-empty attachment')
    if len(data) > MAX_ATTACHMENT_SIZE:
        raise HTTPException(413, 'Attachment size cannot exceed 25 MB')
    filename = Path(file.filename or 'attachment').name[:500]
    content_type = (file.content_type or 'application/octet-stream')[:200]
    with db() as con:
        con.execute('''UPDATE contracts SET attachment_name=?,attachment_content_type=?,attachment_data=?,attachment_size=?,updated_at=? WHERE id=?''',
                    (filename, content_type, data, len(data), datetime.now(timezone.utc).isoformat(), contract_id))
    write_audit(user,'attachment uploaded','contract',contract_id,filename)
    return {'success': True, 'filename': filename, 'size': len(data)}


@app.get('/api/contracts/{contract_id}/attachment')
def download_contract_attachment(contract_id: int, authorization: str | None = Header(None)):
    user = auth(authorization)
    contract_access(user, contract_id)
    with db() as con:
        row = con.execute('SELECT attachment_name,attachment_content_type,attachment_data FROM contracts WHERE id=?',(contract_id,)).fetchone()
    if not row or row['attachment_data'] is None:
        raise HTTPException(404, 'This contract has no attachment')
    filename = row['attachment_name'] or 'attachment'
    disposition = "attachment; filename*=UTF-8''" + quote(filename)
    return StreamingResponse(io.BytesIO(bytes(row['attachment_data'])), media_type=row['attachment_content_type'] or 'application/octet-stream', headers={'Content-Disposition': disposition})


@app.delete('/api/contracts/{contract_id}/attachment')
def delete_contract_attachment(contract_id: int, authorization: str | None = Header(None)):
    user = auth(authorization)
    require_permission(user,'can_edit')
    contract_access(user, contract_id)
    with db() as con:
        con.execute('UPDATE contracts SET attachment_name=NULL,attachment_content_type=NULL,attachment_data=NULL,attachment_size=NULL,updated_at=? WHERE id=?',
                    (datetime.now(timezone.utc).isoformat(), contract_id))
    write_audit(user,'attachment removed','contract',contract_id)
    return {'success': True}


@app.delete('/api/contracts/{contract_id}')
def archive_contract(contract_id:int, authorization: str | None = Header(None)):
    user=auth(authorization); require_permission(user,'can_archive'); contract_access(user,contract_id)
    now=datetime.now(timezone.utc).isoformat()
    with db() as con: con.execute('UPDATE contracts SET archived=1,archived_at=?,archived_by=?,archive_exempt=0,updated_at=?,updated_by=? WHERE id=?',(now,user['username'],now,user['username'],contract_id))
    write_audit(user,'archived','contract',contract_id)
    return {'success':True}


@app.delete('/api/contracts/{contract_id}/permanent')
def permanently_delete_contract(contract_id:int, authorization: str | None = Header(None)):
    user=auth(authorization); require_permission(user,'can_delete'); contract_access(user,contract_id)
    with db() as con:
        row=con.execute('SELECT name,property_name FROM contracts WHERE id=?',(contract_id,)).fetchone()
        if not row:
            raise HTTPException(404,'Contract not found')
        con.execute('DELETE FROM contracts WHERE id=?',(contract_id,))
    write_audit(user,'permanently deleted','contract',contract_id,row['name'],row['property_name'])
    return {'success':True}


@app.post('/api/contracts/{contract_id}/restore')
def restore_contract(contract_id:int, authorization: str | None = Header(None)):
    user=auth(authorization); require_permission(user,'can_restore'); contract=contract_access(user,contract_id); now=datetime.now(timezone.utc).isoformat()
    with db() as con: con.execute('UPDATE contracts SET archived=0,archived_at=NULL,archived_by=NULL,archive_exempt=1,updated_at=?,updated_by=? WHERE id=?',(now,user['username'],contract_id))
    write_audit(user,'restored','contract',contract_id,'Restored from archive',contract['property_name'])
    return {'success':True}


@app.get('/api/audit')
def audit_report(authorization: str | None = Header(None)):
    user=auth(authorization)
    sql='''SELECT TOP 500 id,username,action,entity_type,entity_id,details,property_name,
        CONVERT(NVARCHAR(40),created_at,127) AS created_at FROM audit_logs'''
    args=[]
    if user['property_name']:
        sql+=' WHERE property_name=?'; args.append(user['property_name'])
    sql+=' ORDER BY created_at DESC'
    with db() as con: return [dict(x) for x in con.execute(sql,args)]


@app.get('/api/stats')
def stats(authorization: str | None = Header(None)):
    rows=list_contracts(False,authorization); counts={'total':len(rows),'active':0,'expiring_soon':0,'expired':0,'value':0}
    for r in rows: counts[r['status']]+=1; counts['value']+=r['value']
    return counts


def admin(authorization: str | None) -> dict[str, Any]:
    user=auth(authorization)
    require_permission(user,'can_manage_users')
    return user


@app.get('/api/departments')
def get_departments(authorization: str | None = Header(None)):
    auth(authorization)
    with db() as con: return [dict(x) for x in con.execute('SELECT id,name,description,active FROM departments WHERE active=1 ORDER BY name')]


@app.post('/api/departments')
def add_department(p:NamedPayload, authorization: str | None = Header(None)):
    user=admin(authorization); name=p.name.strip()
    if not name: raise HTTPException(400,'Department name is required')
    now=datetime.now(timezone.utc).isoformat()
    with db() as con:
        if con.execute('SELECT id FROM departments WHERE name=?',(name,)).fetchone(): raise HTTPException(409,'Department already exists')
        cur=con.execute('INSERT INTO departments(name,description,created_at) OUTPUT INSERTED.id VALUES(?,?,?)',(name,p.description.strip(),now)); entity_id=cur.fetchone()['id']
    write_audit(user,'created','department',entity_id,name)
    return {'success':True,'id':entity_id}


@app.put('/api/departments/{department_id}')
def update_department(department_id:int, p:NamedPayload, authorization: str | None = Header(None)):
    user=admin(authorization); name=p.name.strip()
    with db() as con:
        old=con.execute('SELECT name FROM departments WHERE id=?',(department_id,)).fetchone()
        if not old: raise HTTPException(404,'Department not found')
        con.execute('UPDATE departments SET name=?,description=? WHERE id=?',(name,p.description.strip(),department_id)); con.execute('UPDATE contracts SET department=? WHERE department_id=?',(name,department_id))
    write_audit(user,'updated','department',department_id,f"{old['name']} → {name}")
    return {'success':True}


@app.delete('/api/departments/{department_id}')
def delete_department(department_id:int, authorization: str | None = Header(None)):
    user=admin(authorization)
    with db() as con:
        old=con.execute('SELECT name FROM departments WHERE id=?',(department_id,)).fetchone()
        if not old: raise HTTPException(404,'Department not found')
        if con.execute('SELECT COUNT(*) AS count FROM contracts WHERE department_id=? OR department=?',(department_id,old['name'])).fetchone()['count']: raise HTTPException(400,'This department is used by contracts and cannot be deleted')
        con.execute('DELETE FROM departments WHERE id=?',(department_id,))
    write_audit(user,'deleted','department',department_id,old['name'])
    return {'success':True}


@app.get('/api/currencies')
def get_currencies(authorization: str | None = Header(None)):
    auth(authorization)
    with db() as con: return [dict(x) for x in con.execute('SELECT code,name,symbol,active FROM currencies WHERE active=1 ORDER BY code')]


@app.post('/api/currencies')
def add_currency(p:CurrencyPayload, authorization: str | None = Header(None)):
    user=admin(authorization); code=p.code.strip().upper()[:20]
    if not code or not p.name.strip(): raise HTTPException(400,'Currency code and name are required')
    now=datetime.now(timezone.utc).isoformat()
    with db() as con:
        if con.execute('SELECT code FROM currencies WHERE code=?',(code,)).fetchone(): raise HTTPException(409,'Currency already exists')
        con.execute('INSERT INTO currencies(code,name,symbol,created_at) VALUES(?,?,?,?)',(code,p.name.strip(),p.symbol.strip(),now))
    write_audit(user,'created','currency',None,code)
    return {'success':True,'code':code}


@app.put('/api/currencies/{code}')
def update_currency(code:str, p:CurrencyPayload, authorization: str | None = Header(None)):
    user=admin(authorization); new_code=p.code.strip().upper()[:20]
    with db() as con:
        old=con.execute('SELECT code FROM currencies WHERE code=?',(code,)).fetchone()
        if not old: raise HTTPException(404,'Currency not found')
        if new_code!=code and con.execute('SELECT code FROM currencies WHERE code=?',(new_code,)).fetchone(): raise HTTPException(409,'Currency code already exists')
        con.execute('UPDATE currencies SET code=?,name=?,symbol=? WHERE code=?',(new_code,p.name.strip(),p.symbol.strip(),code)); con.execute('UPDATE contracts SET currency=?,currency_code=? WHERE currency_code=? OR currency=?',(new_code,new_code,code,code))
    write_audit(user,'updated','currency',None,f'{code} → {new_code}')
    return {'success':True,'code':new_code}


@app.delete('/api/currencies/{code}')
def delete_currency(code:str, authorization: str | None = Header(None)):
    user=admin(authorization)
    with db() as con:
        if not con.execute('SELECT code FROM currencies WHERE code=?',(code,)).fetchone(): raise HTTPException(404,'Currency not found')
        if con.execute('SELECT COUNT(*) AS count FROM contracts WHERE currency_code=? OR currency=?',(code,code)).fetchone()['count']: raise HTTPException(400,'This currency is used by contracts and cannot be deleted')
        con.execute('DELETE FROM currencies WHERE code=?',(code,))
    write_audit(user,'deleted','currency',None,code)
    return {'success':True}


@app.get('/api/users')
def users(authorization: str | None = Header(None)):
    admin(authorization)
    with db() as con:
        rows=[dict(x) for x in con.execute('''SELECT id,username,role,group_name,property_name,all_properties,
            can_view,can_create,can_edit,can_archive,can_restore,can_delete,can_manage_users
            FROM users ORDER BY username''')]
        for row in rows:
            row['property_names']=[x['property_name'] for x in con.execute('SELECT property_name FROM user_properties WHERE user_id=? ORDER BY property_name',(row['id'],))]
    return rows


def save_user_properties(con: SqlConnection, user_id: int, names: list[str]) -> list[str]:
    cleaned=list(dict.fromkeys(name.strip() for name in names if name.strip()))
    if cleaned:
        placeholders=','.join('?' for _ in cleaned)
        found={x['name'] for x in con.execute(f'SELECT name FROM properties WHERE name IN ({placeholders})',tuple(cleaned))}
        if found != set(cleaned):
            raise HTTPException(400,'One or more selected properties do not exist')
    con.execute('DELETE FROM user_properties WHERE user_id=?',(user_id,))
    for name in cleaned:
        con.execute('INSERT INTO user_properties(user_id,property_name) VALUES(?,?)',(user_id,name))
    return cleaned


@app.post('/api/users')
def add_user(p:UserPayload, authorization: str | None = Header(None)):
    actor=admin(authorization)
    if p.role not in {'admin','it_admin','financial_manager','general_manager','department_head'}: raise HTTPException(400,'Invalid role')
    with db() as con:
        cur=con.execute('''INSERT INTO users(username,password_hash,role,group_name,property_name,all_properties,
            can_view,can_create,can_edit,can_archive,can_restore,can_delete,can_manage_users)
            OUTPUT INSERTED.id VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)''',
            (p.username.strip(),password_hash(p.password or 'Welcome@123'),p.role,p.group_name,p.property_name,
             p.all_properties,p.can_view,p.can_create,p.can_edit,p.can_archive,p.can_restore,p.can_delete,p.can_manage_users)); user_id=cur.fetchone()['id']
        names=save_user_properties(con,user_id,p.property_names or ([p.property_name] if p.property_name else []))
    write_audit(actor,'created','user',user_id,json.dumps({'username':p.username,'role':p.role,'properties':names,'all_properties':p.all_properties}),p.property_name)
    return {'success':True}


@app.put('/api/users/{user_id}')
def update_user(user_id:int, p:UserPayload, authorization: str | None = Header(None)):
    actor=admin(authorization)
    if p.role not in {'admin','it_admin','financial_manager','general_manager','department_head'}: raise HTTPException(400,'Invalid role')
    username=p.username.strip()
    if not username: raise HTTPException(400,'Username is required')
    try:
        with db() as con:
            existing=con.execute('SELECT id,username,role,can_manage_users FROM users WHERE id=?',(user_id,)).fetchone()
            if not existing: raise HTTPException(404,'User not found')
            duplicate=con.execute('SELECT id FROM users WHERE username=? AND id<>?',(username,user_id)).fetchone()
            if duplicate: raise HTTPException(409,'This username is already used by another account')
            if existing['can_manage_users'] and not p.can_manage_users and con.execute('SELECT COUNT(*) AS count FROM users WHERE can_manage_users=1').fetchone()['count']<=1: raise HTTPException(400,'The last user administrator cannot lose user-management permission')
            values=(username,p.role,p.group_name.strip(),p.property_name.strip(),int(p.all_properties),int(p.can_view),int(p.can_create),int(p.can_edit),int(p.can_archive),int(p.can_restore),int(p.can_delete),int(p.can_manage_users))
            if p.password:
                con.execute('''UPDATE users SET username=?,role=?,group_name=?,property_name=?,all_properties=?,can_view=?,can_create=?,can_edit=?,can_archive=?,can_restore=?,can_delete=?,can_manage_users=?,password_hash=? WHERE id=?''',values+(password_hash(p.password),user_id))
            else:
                con.execute('''UPDATE users SET username=?,role=?,group_name=?,property_name=?,all_properties=?,can_view=?,can_create=?,can_edit=?,can_archive=?,can_restore=?,can_delete=?,can_manage_users=? WHERE id=?''',values+(user_id,))
            names=save_user_properties(con,user_id,p.property_names or ([p.property_name] if p.property_name else []))
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(400,f'Could not update user: {exc}') from exc
    for session in SESSIONS.values():
        if session['username']==existing['username']: session['username']=p.username
    write_audit(actor,'updated','user',user_id,json.dumps({'username':p.username,'role':p.role,'properties':names,'all_properties':p.all_properties,'permissions':{'view':p.can_view,'create':p.can_create,'edit':p.can_edit,'archive':p.can_archive,'restore':p.can_restore,'delete':p.can_delete,'manage_users':p.can_manage_users}}),p.property_name)
    return {'success':True}


@app.delete('/api/users/{user_id}')
def delete_user(user_id:int, authorization: str | None = Header(None)):
    actor=admin(authorization)
    with db() as con:
        existing=con.execute('SELECT id,username,role,property_name FROM users WHERE id=?',(user_id,)).fetchone()
        if not existing: raise HTTPException(404,'User not found')
        if existing['username']==actor['username']: raise HTTPException(400,'You cannot delete your own signed-in account')
        if existing['role']=='admin' and con.execute("SELECT COUNT(*) AS count FROM users WHERE role='admin'").fetchone()['count']<=1: raise HTTPException(400,'The last administrator cannot be deleted')
        con.execute('DELETE FROM users WHERE id=?',(user_id,))
    for token,session in list(SESSIONS.items()):
        if session['username']==existing['username']: SESSIONS.pop(token,None)
    write_audit(actor,'deleted','user',user_id,existing['username'],existing['property_name'])
    return {'success':True}


@app.get('/api/groups')
def groups(authorization: str | None = Header(None)):
    auth(authorization)
    with db() as con: return [dict(x) for x in con.execute('SELECT * FROM groups_tbl ORDER BY name')]


@app.post('/api/groups')
def add_group(p:NamedPayload, authorization: str | None = Header(None)):
    actor=admin(authorization)
    with db() as con:
        cur=con.execute('INSERT INTO groups_tbl(name,description) OUTPUT INSERTED.id VALUES(?,?)',(p.name,p.description)); group_id=cur.fetchone()['id']
    write_audit(actor,'created','group',group_id,p.name)
    return {'success':True}


@app.put('/api/groups/{group_id}')
def update_group(group_id:int, p:NamedPayload, authorization: str | None = Header(None)):
    actor=admin(authorization)
    with db() as con:
        old=con.execute('SELECT name FROM groups_tbl WHERE id=?',(group_id,)).fetchone()
        if not old: raise HTTPException(404,'Group not found')
        con.execute('UPDATE groups_tbl SET name=?,description=? WHERE id=?',(p.name.strip(),p.description.strip(),group_id)); con.execute('UPDATE users SET group_name=? WHERE group_name=?',(p.name.strip(),old['name']))
    write_audit(actor,'updated','group',group_id,f"{old['name']} → {p.name}")
    return {'success':True}


@app.delete('/api/groups/{group_id}')
def delete_group(group_id:int, authorization: str | None = Header(None)):
    actor=admin(authorization)
    with db() as con:
        old=con.execute('SELECT name FROM groups_tbl WHERE id=?',(group_id,)).fetchone()
        if not old: raise HTTPException(404,'Group not found')
        if con.execute('SELECT COUNT(*) AS count FROM users WHERE group_name=?',(old['name'],)).fetchone()['count']: raise HTTPException(400,'This group is assigned to users and cannot be deleted')
        con.execute('DELETE FROM groups_tbl WHERE id=?',(group_id,))
    write_audit(actor,'deleted','group',group_id,old['name'])
    return {'success':True}


@app.get('/api/properties')
def properties(authorization: str | None = Header(None)):
    auth(authorization)
    with db() as con: return [dict(x) for x in con.execute('''SELECT p.*,
        CASE WHEN l.logo_data IS NULL THEN CAST(0 AS BIT) ELSE CAST(1 AS BIT) END AS has_logo,
        l.logo_name FROM properties p LEFT JOIN property_logos l ON l.property_name=p.name ORDER BY p.name''')]


@app.post('/api/properties')
def add_property(p:NamedPayload, authorization: str | None = Header(None)):
    actor=admin(authorization)
    with db() as con:
        cur=con.execute('INSERT INTO properties(name,logo) OUTPUT INSERTED.id VALUES(?,?)',(p.name,p.logo)); property_id=cur.fetchone()['id']
    write_audit(actor,'created','property',property_id,p.name,p.name)
    return {'success':True}


@app.put('/api/properties/{name}')
def update_property(name:str, p:NamedPayload, authorization: str | None = Header(None)):
    actor=admin(authorization); new_name=p.name.strip()
    with db() as con:
        existing=con.execute('SELECT id,name FROM properties WHERE name=?',(name,)).fetchone()
        if not existing: raise HTTPException(404,'Property not found')
        if new_name!=name and con.execute('SELECT id FROM properties WHERE name=?',(new_name,)).fetchone(): raise HTTPException(409,'Property name already exists')
        con.execute('UPDATE properties SET name=?,logo=? WHERE id=?',(new_name,p.logo,existing['id'])); con.execute('UPDATE contracts SET property_name=? WHERE property_name=?',(new_name,name)); con.execute('UPDATE users SET property_name=? WHERE property_name=?',(new_name,name)); con.execute('UPDATE user_properties SET property_name=? WHERE property_name=?',(new_name,name)); con.execute('UPDATE property_logos SET property_name=? WHERE property_name=?',(new_name,name))
    for session in SESSIONS.values():
        if session['property_name']==name: session['property_name']=new_name
    write_audit(actor,'updated','property',existing['id'],f'{name} → {new_name}',new_name)
    return {'success':True,'name':new_name}


@app.delete('/api/properties/{name}')
def delete_property(name:str, authorization: str | None = Header(None)):
    actor=admin(authorization)
    with db() as con:
        existing=con.execute('SELECT id FROM properties WHERE name=?',(name,)).fetchone()
        if not existing: raise HTTPException(404,'Property not found')
        if con.execute('SELECT COUNT(*) AS count FROM contracts WHERE property_name=?',(name,)).fetchone()['count']: raise HTTPException(400,'This property is used by contracts and cannot be deleted')
        if con.execute('SELECT COUNT(*) AS count FROM users WHERE property_name=?',(name,)).fetchone()['count'] or con.execute('SELECT COUNT(*) AS count FROM user_properties WHERE property_name=?',(name,)).fetchone()['count']: raise HTTPException(400,'This property is assigned to users and cannot be deleted')
        con.execute('DELETE FROM property_logos WHERE property_name=?',(name,)); con.execute('DELETE FROM properties WHERE id=?',(existing['id'],))
    write_audit(actor,'deleted','property',existing['id'],name,name)
    return {'success':True}


@app.post('/api/properties/{name}/logo')
async def upload_property_logo(name:str, file:UploadFile=File(...), authorization: str | None = Header(None)):
    user=admin(authorization)
    with db() as con:
        if not con.execute('SELECT id FROM properties WHERE name=?',(name,)).fetchone(): raise HTTPException(404,'Property not found')
    content_type=(file.content_type or '').lower()
    if content_type not in ('image/png','image/jpeg','image/webp'): raise HTTPException(400,'Logo must be PNG, JPEG, or WebP')
    data=await file.read(5*1024*1024+1)
    if not data or len(data)>5*1024*1024: raise HTTPException(400,'Logo size must be between 1 byte and 5 MB')
    filename=Path(file.filename or 'logo').name[:500]; now=datetime.now(timezone.utc).isoformat()
    with db() as con:
        con.execute('''MERGE property_logos AS target USING (SELECT ? AS property_name) AS source ON target.property_name=source.property_name
            WHEN MATCHED THEN UPDATE SET logo_name=?,content_type=?,logo_data=?,updated_at=?,updated_by=?
            WHEN NOT MATCHED THEN INSERT(property_name,logo_name,content_type,logo_data,updated_at,updated_by) VALUES(?,?,?,?,?,?);''',
            (name,filename,content_type,data,now,user['username'],name,filename,content_type,data,now,user['username']))
    write_audit(user,'logo uploaded','property',None,filename,name)
    return {'success':True,'name':filename}


@app.get('/api/properties/{name}/logo')
def property_logo(name:str, authorization: str | None = Header(None)):
    auth(authorization)
    with db() as con: row=con.execute('SELECT logo_name,content_type,logo_data FROM property_logos WHERE property_name=?',(name,)).fetchone()
    if not row: raise HTTPException(404,'This property has no logo')
    return StreamingResponse(io.BytesIO(bytes(row['logo_data'])),media_type=row['content_type'],headers={'Content-Disposition':"inline; filename*=UTF-8''"+quote(row['logo_name'])})


@app.delete('/api/properties/{name}/logo')
def delete_property_logo(name:str, authorization: str | None = Header(None)):
    user=admin(authorization)
    with db() as con: con.execute('DELETE FROM property_logos WHERE property_name=?',(name,))
    write_audit(user,'logo removed','property',None,'',name)
    return {'success':True}


@app.post('/api/properties/{name}/activate')
def activate_property(name:str, authorization: str | None = Header(None)):
    admin(authorization)
    with db() as con:
        con.execute('UPDATE properties SET active=0'); con.execute('UPDATE properties SET active=1 WHERE name=?',(name,))
    return {'success':True}


@app.get('/api/active-property')
def active_property(authorization: str | None = Header(None)):
    auth(authorization)
    with db() as con: row=con.execute('SELECT TOP 1 * FROM properties WHERE active=1').fetchone()
    return dict(row) if row else {'name':'Maintenance Contracts','logo':''}


BACKUP_TABLES = {
    'properties': True, 'groups_tbl': True, 'users': True, 'departments': True,
    'currencies': False, 'system_settings': False, 'contracts': True,
    'property_logos': True, 'audit_logs': True, 'user_properties': False,
}
BACKUP_DATETIMEOFFSET_COLUMNS = {
    'departments': {'created_at'}, 'currencies': {'created_at'},
    'system_settings': {'updated_at'},
    'contracts': {'created_at','updated_at','archived_at'},
    'property_logos': {'updated_at'}, 'audit_logs': {'created_at'},
}


def backup_encode(value: Any) -> Any:
    if isinstance(value,(bytes,bytearray,memoryview)): return {'__backup_type__':'bytes','value':base64.b64encode(bytes(value)).decode('ascii')}
    if isinstance(value,Decimal): return {'__backup_type__':'decimal','value':str(value)}
    if isinstance(value,(datetime,date)): return {'__backup_type__':'datetime','value':value.isoformat()}
    return value


def backup_decode(value: Any) -> Any:
    if not isinstance(value,dict) or '__backup_type__' not in value: return value
    if value['__backup_type__']=='bytes': return base64.b64decode(value['value'])
    if value['__backup_type__']=='decimal': return Decimal(value['value'])
    if value['__backup_type__']=='datetime': return value['value']
    raise ValueError('Unsupported backup value type')


def build_database_backup(user: dict[str, Any]) -> tuple[bytes,str]:
    exported={}
    with db() as con:
        for table in BACKUP_TABLES:
            columns=[row['name'] for row in con.execute('SELECT name FROM sys.columns WHERE object_id=OBJECT_ID(?) ORDER BY column_id',(table,))]
            datetime_columns=BACKUP_DATETIMEOFFSET_COLUMNS.get(table,set())
            select_clause=','.join(f'CONVERT(NVARCHAR(40),[{column}],127) AS [{column}]' if column in datetime_columns else f'[{column}]' for column in columns)
            exported[table]=[{key:backup_encode(value) for key,value in row.items()} for row in con.execute(f'SELECT {select_clause} FROM [{table}]')]
    payload={'format':'maintenance-contract-backup','version':1,'created_at':datetime.now(timezone.utc).isoformat(),'created_by':user['username'],'tables':exported}
    manifest=json.dumps(payload,ensure_ascii=False,separators=(',',':')).encode('utf-8'); stream=io.BytesIO()
    with zipfile.ZipFile(stream,'w',zipfile.ZIP_DEFLATED,compresslevel=6) as archive: archive.writestr('backup.json',manifest)
    filename='maintenance-contract-backup-'+datetime.now().strftime('%Y%m%d-%H%M%S')+'.mcbak'
    return stream.getvalue(),filename


@app.get('/api/database/backup')
def database_backup(authorization: str | None = Header(None)):
    user=admin(authorization); data,filename=build_database_backup(user)
    write_audit(user,'database backup created','database',None,filename)
    return StreamingResponse(io.BytesIO(data),media_type='application/octet-stream',headers={'Content-Disposition':f'attachment; filename="{filename}"'})


def apply_database_backup(raw: bytes, user: dict[str, Any], filename: str) -> None:
    if not raw or len(raw)>100*1024*1024: raise HTTPException(400,'Backup file must be smaller than 100 MB')
    try:
        with zipfile.ZipFile(io.BytesIO(raw),'r') as archive:
            info=archive.getinfo('backup.json')
            if info.file_size>300*1024*1024: raise ValueError('Expanded backup is too large')
            payload=json.loads(archive.read(info).decode('utf-8'))
        if payload.get('format')!='maintenance-contract-backup' or payload.get('version')!=1: raise ValueError('Unsupported backup format')
        tables=payload.get('tables',{})
        tables.setdefault('user_properties',[])
        if not all(table in tables and isinstance(tables[table],list) for table in BACKUP_TABLES): raise ValueError('Backup is incomplete')
    except Exception as exc: raise HTTPException(400,f'Invalid backup file: {exc}')
    delete_order=['audit_logs','property_logos','contracts','user_properties','users','groups_tbl','properties','departments','currencies','system_settings']
    insert_order=['properties','groups_tbl','users','user_properties','departments','currencies','system_settings','contracts','property_logos','audit_logs']
    try:
        with db() as con:
            schema={table:{row['name'] for row in con.execute('SELECT name FROM sys.columns WHERE object_id=OBJECT_ID(?)',(table,))} for table in BACKUP_TABLES}
            for table in delete_order: con.execute(f'DELETE FROM [{table}]')
            for table in insert_order:
                rows=tables[table]
                if not rows: continue
                columns=[column for column in rows[0] if column in schema[table]]
                if not columns: continue
                if BACKUP_TABLES[table]: con.execute(f'SET IDENTITY_INSERT [{table}] ON')
                placeholders=','.join('?' for _ in columns); names=','.join(f'[{column}]' for column in columns)
                for row in rows: con.execute(f'INSERT INTO [{table}]({names}) VALUES({placeholders})',tuple(backup_decode(row.get(column)) for column in columns))
                if BACKUP_TABLES[table]: con.execute(f'SET IDENTITY_INSERT [{table}] OFF')
            con.execute("UPDATE users SET all_properties=1,can_view=1,can_create=1,can_edit=1,can_archive=1,can_restore=1,can_delete=1,can_manage_users=1 WHERE role='admin'")
            con.execute("INSERT INTO user_properties(user_id,property_name) SELECT id,property_name FROM users u WHERE property_name<>'' AND NOT EXISTS(SELECT 1 FROM user_properties up WHERE up.user_id=u.id AND up.property_name=u.property_name)")
    except Exception as exc: raise HTTPException(400,f'Database restore failed: {exc}')
    write_audit(user,'database restored','database',None,filename)


@app.post('/api/database/restore')
async def database_restore(file:UploadFile=File(...), authorization: str | None = Header(None)):
    user=admin(authorization); raw=await file.read(100*1024*1024+1)
    apply_database_backup(raw,user,Path(file.filename or 'backup.mcbak').name)
    return {'success':True,'message':'Database restored successfully'}


def configured_backup_directory() -> Path:
    with db() as con: row=con.execute("SELECT setting_value FROM system_settings WHERE setting_key='backup_directory'").fetchone()
    directory=Path(row['setting_value'] if row else str(ROOT/'backups')).expanduser()
    if not directory.is_absolute(): directory=(ROOT/directory).resolve()
    return directory


@app.get('/api/database/backup-settings')
def get_backup_settings(authorization: str | None = Header(None)):
    admin(authorization); return {'backup_directory':str(configured_backup_directory())}


@app.put('/api/database/backup-settings')
def update_backup_settings(payload:BackupSettingsPayload, authorization: str | None = Header(None)):
    user=admin(authorization); directory=Path(payload.backup_directory.strip()).expanduser()
    if not directory.is_absolute(): raise HTTPException(400,'Enter an absolute path on the server, such as C:\\MaintenanceBackups')
    try: directory.mkdir(parents=True,exist_ok=True)
    except Exception as exc: raise HTTPException(400,f'The server cannot create or access this path: {exc}')
    now=datetime.now(timezone.utc).isoformat()
    with db() as con: con.execute("UPDATE system_settings SET setting_value=?,updated_at=? WHERE setting_key='backup_directory'",(str(directory),now))
    write_audit(user,'backup path updated','system_settings',None,str(directory))
    return {'success':True,'backup_directory':str(directory)}


@app.post('/api/database/backups')
def create_stored_backup(authorization: str | None = Header(None)):
    user=admin(authorization)
    try: data,filename=build_database_backup(user)
    except Exception as exc: raise HTTPException(500,f'Could not prepare database backup: {exc}')
    now=datetime.now(timezone.utc).isoformat(); directory=configured_backup_directory(); target=directory/filename
    try:
        directory.mkdir(parents=True,exist_ok=True); target.write_bytes(data)
        with db() as con:
            cur=con.execute('INSERT INTO database_backups(backup_name,backup_data,backup_size,backup_path,created_at,created_by) OUTPUT INSERTED.id VALUES(?,NULL,?,?,?,?)',(filename,len(data),str(target),now,user['username'])); backup_id=cur.fetchone()['id']
    except Exception as exc: raise HTTPException(500,f'Could not store database backup at {target}: {exc}')
    write_audit(user,'stored backup created','database_backup',backup_id,filename)
    return {'success':True,'id':backup_id,'name':filename,'size':len(data),'path':str(target),'created_at':now}


def stored_backup_bytes(backup_id:int) -> tuple[dict[str,Any],bytes]:
    with db() as con:
        row=con.execute('SELECT backup_name,backup_data,backup_path FROM database_backups WHERE id=?',(backup_id,)).fetchone()
        if not row: raise HTTPException(404,'Backup not found')
        if row['backup_path'] and Path(row['backup_path']).is_file(): return row,Path(row['backup_path']).read_bytes()
        if row['backup_data'] is not None: return row,bytes(row['backup_data'])
        chunks=[bytes(x['chunk_data']) for x in con.execute('SELECT chunk_data FROM database_backup_chunks WHERE backup_id=? ORDER BY chunk_index',(backup_id,))]
    if not chunks: raise HTTPException(500,'The saved backup contains no data')
    return row,b''.join(chunks)


@app.get('/api/database/backups')
def list_stored_backups(authorization: str | None = Header(None)):
    admin(authorization)
    with db() as con: return [dict(x) for x in con.execute('''SELECT id,backup_name,backup_size,backup_path,created_by,
        CONVERT(NVARCHAR(40),created_at,127) AS created_at FROM database_backups ORDER BY created_at DESC''')]


@app.get('/api/database/backups/{backup_id}/download')
def download_stored_backup(backup_id:int, authorization: str | None = Header(None)):
    admin(authorization)
    row,data=stored_backup_bytes(backup_id)
    return StreamingResponse(io.BytesIO(data),media_type='application/octet-stream',headers={'Content-Disposition':f'attachment; filename="{row["backup_name"]}"'})


@app.post('/api/database/backups/{backup_id}/restore')
def restore_stored_backup(backup_id:int, authorization: str | None = Header(None)):
    user=admin(authorization)
    row,data=stored_backup_bytes(backup_id); apply_database_backup(data,user,row['backup_name'])
    return {'success':True,'message':'Stored backup restored successfully'}


@app.delete('/api/database/backups/{backup_id}')
def delete_stored_backup(backup_id:int, authorization: str | None = Header(None)):
    user=admin(authorization)
    with db() as con:
        row=con.execute('SELECT backup_name,backup_path FROM database_backups WHERE id=?',(backup_id,)).fetchone()
        if not row: raise HTTPException(404,'Backup not found')
        con.execute('DELETE FROM database_backup_chunks WHERE backup_id=?',(backup_id,))
        con.execute('DELETE FROM database_backups WHERE id=?',(backup_id,))
    if row['backup_path']:
        try: Path(row['backup_path']).unlink(missing_ok=True)
        except OSError: pass
    write_audit(user,'stored backup deleted','database_backup',backup_id,row['backup_name'])
    return {'success':True}


@app.get('/api/export')
def export_csv(authorization: str | None = Header(None)):
    rows=list_contracts(False,authorization); output=io.StringIO(); fields=list(Contract.model_fields)
    writer=csv.DictWriter(output,fieldnames=fields); writer.writeheader(); writer.writerows([{k:r.get(k,'') for k in fields} for r in rows])
    return StreamingResponse(iter([output.getvalue()]),media_type='text/csv',headers={'Content-Disposition':'attachment; filename=contracts.csv'})


@app.post('/api/reports/export-xlsx')
def export_report_xlsx(payload: ReportExportPayload, authorization: str | None = Header(None)):
    auth(authorization)
    if not payload.rows: raise HTTPException(400,'The report has no data to export')
    if len(payload.rows)>5000: raise HTTPException(400,'A report export cannot exceed 5,000 rows')
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as ExcelImage
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    wb=Workbook(); ws=wb.active; ws.title='Report'
    columns=list(payload.rows[0])[:50]; last=max(1,len(columns)); navy='173D50'; teal='1E8F7A'; light='EAF2F4'; white='FFFFFF'
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=last); ws.cell(1,1,payload.title[:200]); ws.cell(1,1).font=Font(size=18,bold=True,color=white); ws.cell(1,1).fill=PatternFill('solid',fgColor=navy); ws.cell(1,1).alignment=Alignment(horizontal='left',vertical='center'); ws.row_dimensions[1].height=46
    ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=last); ws.cell(2,1,payload.filters[:1000]); ws.cell(2,1).font=Font(size=10,color='526C79'); ws.cell(2,1).alignment=Alignment(wrap_text=True); ws.row_dimensions[2].height=25
    ws.merge_cells(start_row=3,start_column=1,end_row=3,end_column=last); ws.cell(3,1,'Generated: '+datetime.now().astimezone().strftime('%Y-%m-%d %H:%M')); ws.cell(3,1).font=Font(size=9,italic=True,color='526C79')
    thin=Side(style='thin',color='D8E3E9')
    for col,name in enumerate(columns,1):
        cell=ws.cell(5,col,name); cell.font=Font(bold=True,color=white); cell.fill=PatternFill('solid',fgColor=teal); cell.alignment=Alignment(horizontal='center'); cell.border=Border(bottom=thin)
    for row_index,row in enumerate(payload.rows,6):
        for col_index,name in enumerate(columns,1):
            value=row.get(name,'')
            if isinstance(value,str) and value[:1] in ('=','+','-','@'): value="'"+value
            cell=ws.cell(row_index,col_index,value); cell.border=Border(bottom=thin); cell.alignment=Alignment(vertical='top',wrap_text=True)
            if row_index%2==0: cell.fill=PatternFill('solid',fgColor='F7FAFB')
            if isinstance(value,(int,float)): cell.number_format='#,##0.00'
    for col_index,name in enumerate(columns,1):
        sample=[str(name)]+[str(row.get(name,'')) for row in payload.rows[:100]]; ws.column_dimensions[get_column_letter(col_index)].width=min(42,max(12,max(len(x) for x in sample)+2))
    ws.freeze_panes='A6'; ws.auto_filter.ref=f'A5:{get_column_letter(last)}{5+len(payload.rows)}'; ws.sheet_view.showGridLines=False; ws.print_title_rows='1:5'; ws.page_setup.orientation='landscape'; ws.page_setup.fitToWidth=1; ws.sheet_properties.pageSetUpPr.fitToPage=True
    if payload.property_name:
        with db() as con: logo=con.execute('SELECT logo_data FROM property_logos WHERE property_name=?',(payload.property_name,)).fetchone()
        if logo:
            try:
                picture=ExcelImage(io.BytesIO(bytes(logo['logo_data']))); picture.width=90; picture.height=40; picture.anchor=f'{get_column_letter(last)}1'; ws.add_image(picture)
            except Exception: pass
    stream=io.BytesIO(); wb.save(stream); stream.seek(0)
    filename='management-report-'+datetime.now().strftime('%Y%m%d-%H%M')+'.xlsx'
    return StreamingResponse(stream,media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',headers={'Content-Disposition':f'attachment; filename="{filename}"'})


@app.post('/api/import')
async def import_csv(file:UploadFile=File(...), authorization: str | None = Header(None)):
    auth(authorization); text=(await file.read()).decode('utf-8-sig'); count=0
    for row in csv.DictReader(io.StringIO(text)):
        add_contract(Contract(**row),authorization); count+=1
    return {'success':True,'imported':count}


app.mount('/static',StaticFiles(directory=ROOT/'static'),name='static')

@app.get('/')
def home():
    return FileResponse(
        ROOT/'static/index.html',
        headers={'Cache-Control':'no-store, no-cache, must-revalidate, max-age=0','Pragma':'no-cache'},
    )
